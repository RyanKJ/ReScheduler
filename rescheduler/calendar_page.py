"""
Module for the interactive calendar page
"""

import Tkinter as tk
import ttk
import tkFileDialog
import tkMessageBox
import getpass
import datetime
import calendar
import bisect
import collections
import re
from employee_page import (EmployeePage, EmployeeList, DepartmentList, 
                           EmployeeInfoForm, EmployeeRepeatUnavailable, 
                           EmployeeVacations)
from sales_page import SalesPage
from datetime_widgets import TimeEntry, DateEntry, yearify
from orm_models import *

from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from openpyxl import Workbook, load_workbook, cell
from openpyxl.styles import Alignment

    
wb = load_workbook('CalendarTemplate.xlsx')
ws = wb['Calendar']


class ReScheduler(object):
    """Create a tab button navigator for user to navigate pages.
    
    This class instantiates several pages to display so that the user can
    browse different pages of the program.
    """
    
    def __init__(self, parent, session):
        """Initiate tabs for user to browse between GUI pages in program.
    
        There are three pages that this init method creates. The first is the
        calendar page which displays the calendar, schedule editor and costs
        of calendars relative to average total monthly sales. The second page
        is the employee/department page where employees are 
        added/edited/removed, and a monthly sales page where total monthly sales
        can be added and removed by the user.
        
        
        Args:
            parent: A parent tkinter frame object.
            session: An sqlalchemy session object using sqlite3.
        """
        
        n = ttk.Notebook(parent)
        n.pack()
        
        dep_list = [d.name for d in session.query(Department).all()]
        now = datetime.datetime.now()
        month, year = int(now.strftime("%m")), int(now.strftime("%Y"))
        date = datetime.date(year, month, 1)
        
        # Calendar page widgets
        self.calendar_frame = ttk.Frame(n)
        self.calendar = CalendarPage(self.calendar_frame, 
                                     session, 
                                     date,
                                     dep_list)
        # Employee page widgets                              
        self.employee_page_frame = ttk.Frame(n)
        self.employee_page = EmployeePage(self.employee_page_frame, 
                                          session)
        self.employee_page.pack()
        # Sales page widgets
        self.sales_page_frame = ttk.Frame(n)
        self.sales_page = SalesPage(self.sales_page_frame, 
                                    session, 
                                    self.calendar)
                                    
        n.add(self.calendar_frame, text="Calendar")
        n.add(self.employee_page_frame, text="Employees And Departments")
        n.add(self.sales_page_frame, text="Monthly Revenue Data")
                                            
                
        
class CalendarPage(tk.Frame):
    """Container page for all collection of widgets for the calendar.
    
    The calendar page consists of 4 collections of widgets: the calendar_menu
    that instantiates the calendar via user selection and also exports the
    current displayed calendar to excel. The calendar_display which displays
    an interactive calendar. The schedule_editor that allows the user to
    add/edit/remove schedules. And a calendar_calc that calculates the net cost
    of schedules for that calendar relative to the average monthly revenue
    for that month.
    
    Attributes:
        dep_list: A string list of all department names.
        side_info_frame: tk.Frame that contains schedule_editor and 
            calendar_calc.
        calendar_menu: Collection of widgets for selecting/saving calendar.
        calendar_display: Collection of widgets for interactive calendar.
        calendar_calc: Collection of widgets to display cost of calendar
            relative to average monthly revenue.
    """
    
    def __init__(self, master, session, date, dep_list):
        """Inits CalendarPage with a date and a department list.
        
        Upon the opening of the program, the init method instantiates the
        relevant collection of widgets. The CalendarPage acts as a quasi 
        controller between the various collection of widgets on the calendar
        page.
        
        Args: 
            master: tk.Frame to contain all child widgets.
            date: datetime.date object containing present month and year.
            dep_list: A string list of all department variables.
        """
        self.session = session
        self.dep_list = dep_list
        self.side_info_frame = tk.Frame(master)
        self.schedule_editor = ScheduleEditor(self.side_info_frame, self)
    
        self.calendar_menu = CalendarMenu(master, self, date, dep_list)
        self.calendar_display = CalendarDisplay(master, self,
                                                self.schedule_editor, 
                                                date, dep_list[0])
                    
        self.side_info_frame.pack(side=tk.LEFT, fill="both", expand=True)                                           
        self.calendar_calc = CalendarCalculator(self.side_info_frame, 
                                                session,
                                                self.calendar_display,
                                                dep_list)

                                                
    def update_costs(self):
        """Update the renvenue calculator widgets."""
        self.calendar_calc.update_costs()
                
                
    def create_calendar(self, dep, date):
        """Call calendar_display with a department/date to create calendar.
        
        Args:
            dep: String object to determine which department for the calendar.
            date: datetime.date object to determine month/year for calendar.
        """
        self.calendar_display.create_calendar(dep, date)
        
        
    def save_calendar_to_excel(self, version):
        """Call calendar_display to save calendar to an excel template.
        
        Args:
            version: String object to determine what version label to call 
                current calendar when exporting to excel.
        """
        self.calendar_display.save_calendar_to_excel(version)
        
        
    def autofill(self):
        """Call calendar_display to execute autofill method."""
        self.calendar_display.autofill()
        
                   
                                            
class CalendarMenu(tk.Frame):
    """Widget collection for instantiating/saving calendars.
    
    The calendar menu has several drop down menus to select month, year,
    and department variables which are then used to instantiate a calendar
    with the calendar_display object. The calendar_menu also features a save
    function into excel and an autofill button that fills any schedule that
    does not have an employee assigned with the 'most eligable' employee.
    (See get_eligables method in EligableModel for explanation of 'eligability'
    and how employees are sorted by eligability.)
    
    Attributes:
        MONTH_TO_INT: dict for converting string month names to int.
        master: tk.Frame parent for all child widgets.
        c_page: calendar_page, a quasi-controller object.
        dep_list: String list of all departments.
        department_var: tk.StringVar for representing selected department.
        month_var: tk.StringVar for representing selected month.
        year_var: tk.StringVar for representing selected year.
        version_var: tk.StringVar for representing selected version.
    """

    MONTH_TO_INT = {"January":1, "February":2, "March":3, "April":4, 
                    "May":5, "June":6, "July":7, "August":8, "September":9, 
                    "October":10, "November":11, "December":12}

    def __init__(self, master, c_page, date, dep_list):
    
        """Initialize calendar selection and saving widgets.
        
        Args:
            master: tk.Frame container for child widgets.
            c_page: Quasi-controller class to call other widget collections.
            date: datetime.date object containing present month and year.
            dep_list: List of strings for department names.
        """
        
        self.master = master
        self.c_page = c_page      
        self.dep_list = dep_list

        # tk.Frame to hold the canvas which will display the y-axis scrollbar
        calendar_menu_frame = tk.Frame(self.master)
        calendar_menu_frame.pack(side="top", fill="both", expand=True)
        
        # Department selection widgets
        department_label = ttk.Label(calendar_menu_frame, 
                                         text = "Department: ")
        department_label.grid(row=0, column=0)
        self.department_var = tk.StringVar(calendar_menu_frame)
        self.department_var.set(self.dep_list[0])
        dep_cb = ttk.Combobox(calendar_menu_frame, 
                              textvariable=self.department_var,
                              values=self.dep_list,
                              width=12,
                              state='readonly')
        dep_cb.grid(row=0, column=1)
        # Month selection widgets
        month_label = ttk.Label(calendar_menu_frame, 
                                    text = "Month: ")
        month_label.grid(row=0, column=2)
        self.month_var = tk.StringVar(calendar_menu_frame)
        self.month_var.set(calendar.month_name[date.month])
        MONTHS = ("January", "February", "March", "April",
                  "May", "June", "July", "August", "September", 
                  "October", "November", "December")
        month_cb = ttk.Combobox(calendar_menu_frame, 
                                textvariable=self.month_var,
                                values=MONTHS,
                                width=12,
                                state='readonly')
        month_cb.grid(row=0, column=3)
        # Year selection widgets
        year_label = ttk.Label(calendar_menu_frame, 
                                   text = "Year: ")
        year_label.grid(row=0, column=4)
        self.year_var = tk.StringVar(calendar_menu_frame)
        self.year_var.set(date.year)
        
        year_list = yearify(date.year, 8)
        
        year_cb = ttk.Combobox(calendar_menu_frame, 
                               textvariable=self.year_var,
                               values=year_list,
                               width=5,
                               state='readonly')
        year_cb.grid(row=0, column=5)
        # Widgets to get calendar with selected year, month, department
        get_calendar_button = ttk.Button(calendar_menu_frame, 
                                             text='Get Calendar', 
                                             command=self.create_cal_click)
        get_calendar_button.grid(row=0, column=6, padx=5)
        # Widgets to export calendar to Excel
        spacing_frame_1 = tk.Frame(calendar_menu_frame)
        spacing_frame_1.grid(row=0, column=7, padx=28)
        version_label = ttk.Label(calendar_menu_frame, 
                                         text = "Export As Version: ")
        version_label.grid(row=0, column=8)
        self.version_var = tk.StringVar(calendar_menu_frame)
        self.version_var.set('A')
        version_cb = ttk.Combobox(calendar_menu_frame, 
                                  textvariable=self.version_var,
                                  values=('A', 'B', 'C', 'D', 'E'),
                                  width=5,
                                  state='readonly')
        version_cb.grid(row=0, column=9)
        
        export_button = ttk.Button(calendar_menu_frame, 
                                       text='Save to Excel', 
                                       command=self.save_calendar_to_excel)
        export_button.grid(row=0, column=10)

        # Widgets for auto-fill schedules without employees
        spacing_frame_2 = tk.Frame(calendar_menu_frame)
        spacing_frame_2.grid(row=0, column=11, padx=28)
        autofill_button = ttk.Button(calendar_menu_frame, 
                                         text='Autofill Schedules', 
                                         command=self.autofill)
        autofill_button.grid(row=0, column=12)
        
        sep_bottom = ttk.Separator(calendar_menu_frame, orient=tk.HORIZONTAL)
        sep_bottom.grid(row=1, column=0, columnspan=13, sticky="ew")
        
        
    def create_cal_click(self):
        """Get date and department variables and call create calendar method."""
        dep = self.department_var.get()
        month = self.MONTH_TO_INT[self.month_var.get()]
        year = int(self.year_var.get())
        date = datetime.date(year, month, 1)
        self.c_page.create_calendar(dep, date)
        
        
    def save_calendar_to_excel(self):
        """Call save to excel method."""
        version = self.version_var.get()
        self.c_page.save_calendar_to_excel(version)
        
        
    def autofill(self):
        """Call autofill method."""
        self.c_page.autofill()
        
        
		
class CalendarDisplay(tk.Frame):
    """Create day and title widgets responsible for displaying calendar.
    
    Calendar_Page is responsible for instantiating the day widgets that 
    display the schedules for each day and the calendar weekday 
    header. (i.e. Sunday, Monday, ...)

    Attributes:
        parent: tk.Frame container for child widgets.
        c_page: Quasi-controller class to call other widget collections.
        schedule_editor: UI for user to add schedules to calendar.
        dep: String for department name of current selected calendar.
        date: datetime.date for current selected calendar.
        day_vc_list: A list of day_vc objects that display corresponding
            information about that day: day number and schedules.
        current_clicked_day: Current active day for user interaction.
        canvas: tk.Canvas container to display scrollbars.
        calendar_frame: tk.Frame container for the day_vc objects.
        calendar_title: tk.Frame container for day headers, ie Sunday, Tuesday.
        DAYS: List to map integers to string representation.
    """
    
    DAYS = ['Sunday', 'Monday', 'Tuesday', 'Wednesday',
            'Thursday', 'Friday', 'Saturday']

    def __init__(self, parent, c_page, schedule_editor, date, dep):
        """Initialize widgets for interactive calendar.
        
        Initializes the various container widgets such as canvas and its 
        scrollbars. Then calls create_calendar method with the supplied date
        and department.
        
        Args:
            parent: A parent tk.Frame object.
            c_page: Quasi-controller class to call other widget collections.
            schedule_editor: UI for user to add schedules to calendar.
            date: datetime.date for current selected calendar.
            dep: String for department name of current selected calendar.
        """
        
        self.parent = parent
        self.c_page = c_page
        self.schedule_editor = schedule_editor
       
        self.dep = dep
        self.date = date
        self.day_vc_list = []
        self.current_clicked_day = None 
        
        # Container and scrollbar widgets
        calendar_holder = tk.Frame(self.parent)
        calendar_holder.pack(side="left")
        self.canvas = tk.Canvas(calendar_holder, 
                                borderwidth=0)
        self.calendar_frame = tk.Frame(self.canvas)
        scrollbar = ttk.Scrollbar(calendar_holder, 
                                     orient="vertical", 
                                     command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right",fill="y")
        self.canvas.pack(side="left")
        self.canvas.create_window((0,0), 
                                  window=self.calendar_frame, 
                                  anchor=tk.NW)
        self.calendar_frame.bind("<Configure>", self.onFrameConfigure)

        # Widget for calendar title
        self.calendar_title = ttk.Label(self.calendar_frame, text="")
        self.calendar_title.grid(row=0, column=0, columnspan=7)

        self.create_calendar(dep, date)
        
        
    def onFrameConfigure(self, event):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"), 
                              width=1060, height=1200)
        	
    
    def click_reset(self):
        """Display all schedules for each day then select first day."""
        self.set_current_clicked_day(self.day_vc_list[0])
        for d in reversed(self.day_vc_list):
            d.set_to_clicked("<button-1>")
            
        
    def set_current_clicked_day(self, day):
        """Set current day_vc for calendar_display to be supplied day_vc.
        
        Args:
            day: day_vc object to be set as the current_clicked_day state.
        """
        self.current_clicked_day = day
		
        
    def get_cal_array(self, year, month):
        """Get a python calendar with Sunday set as first weekeday.
        
        Args:
            year: Int representing year.
            month: Int in range 1-12.
        Returns:
            An array of a given month where each week is represented by another
            array whose elements are the day numbers of the month. The 
            calendar is set so that the weekday starts on a Sunday instead of 
            the default Monday.
        """
        
        sundayCalendar = calendar
        sundayCalendar.setfirstweekday(6)
        return sundayCalendar.monthcalendar(year, month)
        
        
    def create_calendar(self, department, date):
        """Create a calendar for a given date and department.
        
        Creates a calendar with interactable widgets that display schedules
        for a given month, year, and department. e.g. August 2016 calendar
        for drivers. Each day in the calendar is clickable and will display
        schedules for that day.
        
        When a day in the calendar is clicked it is highlighted, which then
        schedules can be added or removed in the schedule editor.
        
        Args:
            department: String object to determine which department for the 
                calendar.
            date: datetime.date object to determine month/year for calendar.
        """
        
        self.clear_calendar()
        self.date = date
        self.dep = department
        calendar_array = self.get_cal_array(self.date.year, self.date.month)
        title = "%s Calendar For %s, %s" % (self.dep, 
                                            calendar.month_name[self.date.month],
                                            str(self.date.year))
        self.calendar_title.config(text=title)
        # Create weekday name column titles, ie Sunday, Monday, Tuesday...
        for i in range(0,7):
            day_header = ttk.Label(self.calendar_frame, 
                                  text=self.DAYS[i],  
                                  borderwidth=1)
            day_header.grid(row=1, column=i)
        # i represents the weeks for that calendar
        # j represents the day of week
        for i in range(0, len(calendar_array)):
            for j in range(0,7):
                day_number = calendar_array[i][j]
                date = None
                if day_number != 0:
                    date = datetime.date(self.date.year, self.date.month, 
                                         day_number)
                day_model = DayModel(self.c_page.session, self, 
                                     date, i, j, self.dep)
                day_vc = DayViewController(self.calendar_frame, self, 
                                           self.schedule_editor, day_model)
                self.day_vc_list.append(day_vc)
        # Display schedules and click first day of that month
        self.click_reset()
        
        
    def clear_calendar(self):
        """Destroy day widgets and clear list for that calendar's days."""
        for d in self.day_vc_list:
            d.destroy()
        self.day_vc_list = []
     
          
    def save_calendar_to_excel(self, version):
        """Export the current state of calendar to an excel spreadsheet. 
        
        When the user clicks the save to excel button, the method iterates
        through all day widgets collecting each day's number (i.e. the 1st or
        2nd of the month) and the string that represetns the potential list
        of schedules for that day and maps them to a corresponding excel cell.
        
        Args:
            version: String object to determine what version label to call 
                current calendar when exporting to excel.
        """
        
        ws['A1'] = calendar.month_name[self.date.month]
        ws['B1'] = self.date.year
        ws['C1'] = self.dep
        ws['D1'] = 'Version:  ' + version
        for d in self.day_vc_list:
            coordinates = d.get_excel_coordinates()
            # Header is the cell for the day number, ie 3, 14, or 31
            # Body is the cell for that day's list of schedules
            header, body = coordinates[0], coordinates[1]
            # Write relevant data to excel cells corresponding to this day
            ws[header] = d.day_number
            ws[body] = d.get_text_for_excel()
            ws[body].alignment = Alignment(horizontal='center', 
                                           vertical='center', 
                                           wrapText=True)
        
        filename = '%s-%s%s-ver %s.xlsx' % (self.dep,
                                            calendar.month_name[self.date.month], 
                                            self.date.year,
                                            version_var.get())
                               
        file_opt = {}
        file_opt['defaultextension'] = '.xlsx'
        file_opt['filetypes'] = [('Excel Files', '.xlsx')]
        file_opt['initialdir'] = '/home/' + getpass.getuser() + '/Desktop/'
        file_opt['initialfile'] = filename
        file_opt['parent'] = self.parent
        file_opt['title'] = 'Save Calendar'
        
        dest_filename = tkFileDialog.asksaveasfilename(**file_opt)                                                      
        wb.save(filename = dest_filename)
        
       
    def autofill(self):
        """Fill each unassaigned schedule with most eligable employee
    
        In chronological order of days in calendar, iterate through all 
        schedules and pick the first available (most eligable according to 
        algorithm) employee and set as that schedule's assign employee.
        """
   
        for d in self.day_vc_list:
            schedules = d.schedule_widgets
            for s in schedules:
                s.set_to_clicked("<button-1>")
                # Case where no employee assigned and potential employees exist
                if s.db_schedule.employee_id is None and s.eligable_list != []:
                    s.eligable_listbox.selection_set(0)
                    s.eligable_lb_click('<<ListboxSelect>>')
                    
                    
    def update_costs(self):
        self.c_page.update_costs()
    
    
    
class DayViewController(tk.Frame):
    """Create widgets to display day and update associated day model.
    
    DayViewController is responsible for displaying the relevant data 
    represented in the day model and to also call day model to inform any
    updates via user interaction, hence this class functions to both display
    the model and to pass information on updating day model.

    Attributes:
        DAY_COL_EXCEL: dict to map tkinter grid coordinates to excel 
            coordinates.
        cal: reference to calendar display to inform if self is clicked.
        schedule_editor: reference to display schedule widgets.
        day_model: model representing this day.
        day_number: string representing day number, "" if not of selected 
            month.
        current_clicked_schedule: current schedule selected by user.
        schedule_widgets: dict of db schedule primary keys as keys and schedule
            widget objects as values.
        eligable_vc: dict of db schedule primary keys as keys and eligable_vc
            as values.
        day_frame: tk.Frame container for sub-widgets.
        number_label: tk.Label for day number.
        schedules_lb: tk.Listbox for displaying schedules.
        schedule_display: tk.Frame for displaying parallel list of schedule
            widgets and eligable_vc.
        schedule_frame: tk.Frame for packing schedule widgets.
        eligable_frame: tk.Frame for packing eligable widgets.
    """

    DAY_COL_EXCEL = {0:'A', 1:'B', 2:'C', 3:'D', 4:'E', 5:'F', 6:'G'}

    def __init__(self, master, calendar_display, schedule_editor, day_model):
        """Initialize the day view according to model representation
    
        Create appropriate widgets and fill with information via day_model
        reference.
        
        Args:
            master: tk.Frame container for subwidgets.
            calendar_display: interactive calendar containing day_vc instances.
            schedule_editor: UI interface to add/remove/edit schedules.
            day_model: corresponding model for this particular date.
        """

        self.cal = calendar_display
        self.schedule_editor = schedule_editor
        self.day_model = day_model
        self.day_number = ""
        if self.day_model.date:
             self.day_number = str(self.day_model.date.day)
 
 
        self.current_clicked_schedule = None
        
        self.schedule_widgets = {}
        self.eligable_vc = {}
    
        self.day_frame = tk.Frame(master, borderwidth=1, 
                                  relief=tk.RIDGE, bg="white")               
        coor = self.get_grid_coordinates()
        self.day_frame.grid(row=coor[0], column=coor[1])
        self.number_label = tk.Label(self.day_frame, 
                                     text=self.day_number,
                                     font=('Tahoma', 10, tk.NORMAL),
                                     width=18, height=1, 
                                     anchor=tk.NW, 
                                     bg="white")
        self.number_label.pack()
          
        self.schedules_lb = tk.Listbox(self.day_frame, activestyle='none', 
                                       exportselection=0,
                                       font=('Tahoma', 10, tk.NORMAL),
                                       selectmode=tk.SINGLE, borderwidth=0, 
                                       highlightthickness=0, bg="white")
        self.schedules_lb.pack(padx=5)
        self.number_label.bind('<Button-1>', lambda event: self.set_to_clicked(event))
        self.number_label.bind('<Enter>', lambda event: self.mouse_enter(event))
        self.number_label.bind('<Leave>', lambda event: self.mouse_leave(event))
        self.schedules_lb.bind('<Button-1>', lambda event: self.set_to_clicked(event))
        self.schedules_lb.bind('<<ListboxSelect>>', 
                               lambda event: self.schedule_lb_click(event))
        self.schedules_lb.bind('<Double-1>', 
                               lambda event: self.schedule_lb_click(event))
        self.schedules_lb.bind('<Enter>', 
                               lambda event: self.mouse_enter(event))
        self.schedules_lb.bind('<Leave>', 
                               lambda event: self.mouse_leave(event))
                               
                    
        self.schedule_display = tk.Frame(self.schedule_editor.schedule_widgets_frame)
        self.title_var = tk.StringVar(self.schedule_display)                             
        
        self.schedule_title = tk.Label(self.schedule_display, 
                                       textvariable=self.title_var)
        self.schedule_title.pack(side=tk.TOP)                                         
                                              
        self.schedule_frame = tk.Frame(self.schedule_display)  
        self.schedule_frame.pack()
        self.eligable_frame = ttk.LabelFrame(self.schedule_display,
                                             text="Eligable Employees For Selected Schedule:")
        self.eligable_frame.pack(pady=6)
        if self.day_model.date: 
            title_str = "%s Schedules for %s %s" % (self.day_model.dep,
                                                     calendar.month_name[self.day_model.date.month],
                                                     self.day_number)
            self.title_var.set(title_str)
            self.create_schedules_and_eligable_vc()

        
    def create_schedules_and_eligable_vc(self):
        """Create schedule widgets and eligable_vc for this date."""
        self.reset_sw_and_eligables()
        schedules = self.day_model.schedules
        for id in schedules:
            str = self.day_model.schedule_strings[id]
            schedule_widget = ScheduleWidget(self.schedule_frame, 
                                             id, str, self)
            self.schedule_widgets[id] = schedule_widget
            self.schedules_lb.insert(tk.END, str)
            # Create an eligable_viewcontroller instance for this schedule
            eligable_model = self.day_model.eligable_models[id]
            eligable_vc = EligableViewController(self.eligable_frame,
                                                 eligable_model,
                                                 self)
            self.eligable_vc[id] = eligable_vc                
                                       
    
    def set_to_clicked(self, event):
        """Click this day_vc and display associated schedules in GUI.
        
        Args:
            event: tk event object from a user left mouse click.
        """
        
        if self.day_number != "" and self.cal.current_clicked_day != self:
            self.cal.current_clicked_day.set_to_unclicked()
            self.day_frame.config(bg="LightSkyBlue")
            self.number_label.config(bg="LightSkyBlue")
            self.schedules_lb.config(bg="LightSkyBlue")
            self.schedule_display.pack()
            self.cal.set_current_clicked_day(self)   
            
            
    def set_to_unclicked(self):
        """Set this day_vc to a non-clicked state.

        Non-clicked state means that the day_vc's sub-widget's color is 
        set white from light blue to signal to the user that this day_vc
        is not highlighted and thus no longer an active (or clicked) day_vc 
        to add/remove schedules from.
        """
        
        self.day_frame.config(bg="White")
        self.number_label.config(bg="White")
        self.schedules_lb.config(bg="White")
        self.schedules_lb.selection_clear(0, tk.END)    
        self.schedule_display.pack_forget()
        
        
    def mouse_enter(self, event):
        """Highlights day_widgets to display mouse enter event.
        
        Args:
            event: tk event object from a user mouse over tk object.
        """
        
        if self.day_number != "" and self.cal.current_clicked_day != self:
            self.day_frame.config(bg="#ccebff")
            self.number_label.config(bg="#ccebff")
            self.schedules_lb.config(bg="#ccebff")
        
        
    def mouse_leave(self, event):
        """De-highlights day_widgets to display mouse leave event.
        
        Args:
            event: tk event object from a user mouse over tk object.
        """
        
        if self.day_number != "" and self.cal.current_clicked_day != self:
            self.day_frame.config(bg="white")
            self.number_label.config(bg="white")
            self.schedules_lb.config(bg="white")
         
         
    def reset_sw_and_eligables(self):
        """Destroy schedule widgets and eligable_vc, reset selected schedule."""
        self.schedules_lb.delete(0, tk.END)
        for sw in self.schedule_widgets:
            self.schedule_widgets[sw].destroy()
        for evc in self.eligable_vc:
            self.eligable_vc[evc].destroy()
        self.current_clicked_schedule = None
            
            
    def destroy(self):
        """Delete the widgets that are associated with this day_vc."""
        self.number_label.destroy()
        self.schedules_lb.destroy()
        self.day_frame.destroy()
        self.schedule_display.destroy()
        for s in self.schedule_widgets:
            self.schedule_widgets[s].destroy()

            
    def schedule_widget_click(self, event, schedule_widget):
        """Select schedule via schedule widget and show appropriate widgets.
        
        Args:
            event: tk event object from a user left mouse click on a schedule
                widget.
            schedule_widget: The schedule_widget object that was clicked upon.
        """
        
        id = schedule_widget.pk
        if id != self.current_clicked_schedule:
            schedule_widget.set_to_clicked()
            self.curr_schedule_unclick()
            self.current_clicked_schedule = id
            # highlight corresponding str in parallel listbox
            index = self.day_model.schedules.index(id)
            self.listbox_schedule_highlight(index)
            # Display any potential employees in eligable listbox
            self.eligable_vc[id].show()
        
        
    def schedule_lb_click(self, event):
        """Select a particular schedule in this day_vc's listbox.
        
        Args:
            event: tk event object from a user left mouse click on this day's
                listbox.
        """
        
        cursel = self.schedules_lb.curselection()
        if cursel != ():
            self.curr_schedule_unclick()
            index = cursel[0]
            id = self.day_model.schedules[index]
            self.current_clicked_schedule = id
            self.schedule_widgets[id].set_to_clicked()
            # Display any potential employees in eligable listbox
            self.eligable_vc[id].show()

        
    def curr_schedule_unclick(self):
        """Hide or reset relevant widgets currently clicked."""
        if self.current_clicked_schedule != None:
            # reset stuff here
            id = self.current_clicked_schedule
            schedule_widget = self.schedule_widgets[id]
            schedule_widget.set_to_unclicked()
            eligable_vc = self.eligable_vc[id]
            eligable_vc.clear_values()
            eligable_vc.hide()
            
            
    def highlight_new_schedule(self, id):
        """Refresh view for new added schedule and highlight new schedule.
        
        Args:
            id: The primary key of the schedule to be highlighted in the view.
        """
        
        self.create_schedules_and_eligable_vc()
        
        sw = self.schedule_widgets[id]
        self.schedule_widget_click("<Button-1>", sw)
        
            
    def listbox_schedule_highlight(self, index):
        """Highlight schedule in listbox corresponding to selected schedule.
        
        Args:
            index: The index of the listbox to be highlighted.
        """
        
        self.schedules_lb.selection_clear(0, tk.END)
        self.schedules_lb.selection_set(index)

        
    def update_schedule_str(self, id, new_str):
        """Update the schedule string with given string argument.
        
        Args:
            id: Primary key of the schedule to be updated.
            new_str: The new string to represent the schedule in the view.
        """
        
        self.day_model.schedule_strings[id] = new_str
        
        index = self.day_model.schedules.index(id)
        self.edit_employee_name_text(index, new_str)
        schedule_widget = self.schedule_widgets[id]
        schedule_widget.set_text(new_str)
        
    
    def edit_employee_name_text(self, index, new_str):
        """Replace an employee name for particular schedule with a new name.
        
        Args: 
            index: Index in the listbox that will be edited.
            new_str: The new string to represent the schedule in the view.
        """
        
        self.schedules_lb.delete(index)
        self.schedules_lb.insert(index, new_str)
        self.schedules_lb.selection_set(index)
        
        
    def remove_schedule(self, id):
        """Removes schedule from display then informs model to remove schedule.
        
        The primary key of the db schedule is used as the reference used to
        delete/destroy all relevent display regarding schedule. Then the PK
        is sent to the model and its remove method is called.
        
        Args:
            id: Primary key of the schedule to be removed from the view then
                from model.
        """
        
        del self.schedule_widgets[id]
        index = self.day_model.schedules.index(id)
        self.remove_deleted_schedule_text(index)
        # Case where this instance which is to be deleted 
        # is the schedule that is clicked by the user currently
        if id == self.current_clicked_schedule:
            self.current_clicked_schedule = None

        self.eligable_vc[id].destroy()
        del self.eligable_vc[id]
        
        self.day_model.remove_schedule(id)
        
        
    def remove_deleted_schedule_text(self, index):
        """Delete a schedule text in this day_vc's listbox.
        
        Args:
            index: Index in the listbox that will be edited. 
        """
        
        self.schedules_lb.delete(index)
        
        
    def get_grid_coordinates(self):
        """Get tkinter grid coordinates with respect to the day's date.
        
        Returns:
            A 2-element tuple containing the tkinter grid coordinates for the
            frame containing widgets to display this day.
        """
        
        row = self.day_model.week_number + 2
        col = self.day_model.weekday
        return (row, col)
        
        
    def get_text_for_excel(self):
        """Get all schedules formatted into a string for an excel cell.
        
        Returns:
            A string object containing all the schedule strings for an excel
            spreadsheet cell.
        """
        
        text = ""
        schedules_text = self.schedules_lb.get(0, tk.END)
        
        for s in schedules_text:
            text += s + "\n"
        
        return text    
        
        
    def get_excel_coordinates(self):
        """Get list of strings representing excel coordinates for this date.
    
        Returns:
            Two coordinates are returned from this method. The first is the
            header coordinate for the day number then the second coordinate for
            the body representing this day's schedules.
        """
        
        week_row_header = str((self.day_model.week_number * 2) + 4)
        week_row_body = str((self.day_model.week_number * 2) + 5)
        day_col = self.DAY_COL_EXCEL[self.day_model.date.weekday()]

        return [day_col + week_row_header, day_col + week_row_body]
    
    
      
class DayModel(object):     
    """Represent the day in a month and its schedules
    
    A day is essentially represented by a day number, ie 17th of February.
    A particular department that day represents, the week number relative to 
    that month, ie the 3rd week of the month, and an ordered list of schedules
    ordered by start time from earliest to latest.
    
    Attributes:
        cal: The calendar_display that allows a callback to update costs.
        date: datetime.date object to represent date of this day_model.
        week_number: int of week in the month.
        weekday: int of day in the month.
        dep: string representing the department this day represents.
        schedules: A sorted list of primary keys referencing schedules in db.
        schedule_strings: A dict of schedule primary keys referencing 
            string of db schedules.
        eligable_models: A dict of primary keys referencing eligable model 
            for schedule. 
    """

    def __init__(self, session, calendar_display, date, week_number,
                 weekday, department):
        """Initialize the model of particular day and department
    
        day_model uses the supplied arguments date and dep to fetch appropriate
        database schedules then processes their fields to create a model
        representation of that day. The model consists of the sorted list of 
        primary keys of schedules in the database, a dictionary that maps these
        primary keys to a string to be displayed in the view, and a dictionary
        to map the primary keys to an eligable model that is a sorted list of 
        eligable employees for a particular schedule.
        
        Args:
            calendar_display: calendar containing reference when updating costs
            date: datetime.date object representing date for this day
            week_number: int of week in the month
            weekday: int of day in the month
            dep: string representing the department this day represents
        """
        
        self.session = session
        self.cal = calendar_display
        self.date = date
        self.week_number = week_number
        self.weekday = weekday
        self.dep = department
        
        self.schedules = []
        self.schedule_strings = {}
        self.eligable_models = {}
        
        if date:
            self.get_schedule_id_and_str()
            self.create_eligable_models()
         
        
    def get_schedule_id_and_str(self):
        """Set schedules sorted list and schedule_strings dict"""
        date = datetime.date(self.date.year, self.date.month, 
                             self.date.day)
        db_schedules = (self.session
                            .query(DB_Schedule)
                            .filter(DB_Schedule.schedule_date == date)
                            .all())
                         
        db_schedules = [s for s in db_schedules if s.department == self.dep]
        db_schedules.sort(key=lambda schedule: schedule.start_time)
        
        # List of sorted schedule id's
        self.schedules = [s.id for s in db_schedules]
        # Create a dict of schedule strings with schedule id as key
        for s in db_schedules:
            str = self.get_schedule_str(s)
            self.schedule_strings[s.id] = str
            
            
    def create_eligable_models(self):
        """Create eligable_model for each schedule fetched from database."""
        for id in self.schedules:
            e_model = EligableModel(self.session, id, self.dep, self)
            self.eligable_models[id] = e_model

        
    def get_schedule_str(self, schedule):
        """Get str displaying start and end times and employee if assigned.
        
        Args:
            schedule: A schedule object.
        
        Returns:
            str: A string representing the string version of the schedule to
            be displayed by the view.
        """
        
        start_str, end_str = "", ""
        if schedule.s_undetermined_time:
            start_str = "?"
        else:
            start_str = schedule.start_datetime.strftime("%I:%M")
            
        if schedule.e_undetermined_time:
            end_str = "?"
        else:
            end_str = schedule.end_datetime.strftime("%I:%M")    

        str = start_str + " - " + end_str
        if schedule.employee_id != None:
            employee = (self.session
                            .query(Employees)
                            .filter(Employees.employee_id
                                    == schedule.employee_id)
                            .first())
            str += "  " + employee.first_name
        
        return str    
        
        
    def insert_new_schedule(self, start, end, s_hide, e_hide, dep):
        """Create new db schdule, refresh model, then return id of schedule.
        
        Args:
            start: datetime.datetime object for start datetime of schedule.
            end: datetime.datetime object for end datetime of schedule.
            s_hide: Boolean to determine to hide start time string in the view.
            e_hide: Boolean to determine to hide end time string in the view.
            dep: String to determine department the schedule belongs to.
            
        Returns:
            The primary key of the schedule for reference.
        """
        
        db_schedule = DB_Schedule(start, 
                                  end,
                                  s_hide,
                                  e_hide,
                                  dep)
        self.session.add(db_schedule)
        self.session.commit()
    
        self.reset_values()
        self.get_schedule_id_and_str()
        self.create_eligable_models()
        
        return db_schedule.id
        
        
    def reset_values(self):
        """Clear values of schedule ids, schedule str, and eligable models."""
        self.schedules = []
        self.schedule_strings = {}
        self.eligable_models = {}
        
        
    def remove_schedule(self, id):
        """Remove schedule from model and db.
        
        Args:
            id: The primary key of the schedule to be removed.
        """
        
        db_schedule = (self.session.query(DB_Schedule)
                                   .filter(DB_Schedule.id == id)
                                   .first())
        self.session.delete(db_schedule)
        self.session.commit()
        self.cal.update_costs()
        
        self.schedules.remove(id)
        del self.schedule_strings[id]
        del self.eligable_models[id]
        
        
    def update_costs(self):
        """Call calendar_display to update cost of calendar."""
        self.cal.update_costs()
        
        
        
class EligableViewController(tk.Frame):
    """Create widgets to display sorted list of employees eligable for schedule.
    
    EligableViewController is responsible for displaying the listbox containing
    a sorted list of employees. This list is sorted by several criteria such
    as, but not inclusive to: conflict with overlapping schedules, overtime,
    and vacations.

    Attributes:
        ELIGABLE_WARNING: A dict to map eligability flag to string dialogue
            warning.
        e_model: model to represent sorted list of employee eligability
        day_vc: day_vc shows and hides various eligable_vc according to current
            selected schedule.
        eligable_listbox: tk.Listbox containing sorted list of strings supplied
            by the eligable model.
    """

    ELIGABLE_WARNING = {'(O)': "This employee will be in overtime if assigned to selected schedule, assign anyways?", 
                        '(U)': "This employee has a conflicting repeating unavailability with selected schedule, assign anyways?", 
                        '(V)': "This employee has a conflicting vacation or absent period with selected schedule, assign anyways?", 
                        '(S)': "This employee has a conflicting assigned schedule with selected schedule, assign anyways?"}

    def __init__(self, parent, eligable_model, day_vc):
        """Initialize the eligable view according to model representation.
    
        Create appropriate widgets and fill with information via eligable_model.
        The eligable view is mainly a listbox containing the sorted employees
        with respect to an eligability heuristic score.
        
        Args:
            parent: tk.Frame container for subwidgets.
            eligable_model: model representing sorted list of eligable employees
                for a particular schedule.
            day_vc: view for the day, the reference is used by eligable view to 
                inform day_vc if user edits a schedule by selecting an employee
                to assign.
        """
        
        self.e_model = eligable_model
        self.day_vc = day_vc
        
        self.eligable_listbox = tk.Listbox(parent, 
                                           width=32, height=8, 
                                           exportselection=0,
                                           font=('Tahoma', 11, tk.NORMAL),
                                           bg="white")
        self.eligable_listbox.bind('<<ListboxSelect>>', self.eligable_lb_click)
          
          
    def clear_values(self):
        """Unpack the eligable listbox and clear all strings in listbox."""
        self.eligable_listbox.pack_forget()
        self.eligable_listbox.delete(0, tk.END)
    
    
    def show(self):
        """Clear listbox then reload sorted eligability list and display list."""
        self.clear_values()
        self.display_eligables()
        self.eligable_listbox.pack()
        
    
    def hide(self):
        """Unpack the listbox to hide from user."""
        self.eligable_listbox.pack_forget()
        
        
    def eligable_lb_click(self, event):
        """Assign clicked employee to schedule if user clicks okay on dialog.
        
        This method calls a warning dialog if there is a warning flag. If no
        warning flag or user clicks okay on dialog, the employee is assigned
        to schedule and the model and views are correspondingly updated. If the
        user clicks cancel on warning dialog, the previous state of the listbox
        curselction is set to current state and nothing is changed.
        """
        
        index = self.eligable_listbox.curselection()[0]
        listbox_str = self.eligable_listbox.get(index)                
        if self.eligable_warning(listbox_str):           
            new_str = self.e_model.assign_employee_schedule(index)
            if new_str:
                self.day_vc.update_schedule_str(self.e_model.schedule_pk, new_str)
        else:
            self.click_assigned_employee()
            
            
            
    def eligable_warning(self, employee_str):
        """Create warning dialog if user selects employee with scheduling conflicts.
        
        The warning dialog only triggers if the string representing employee
        name contains a warning flag of the pattern (x) where x is some 
        potential alphanumeric value.
        
        Args:
            employee_str: A string representing employee name for the view.
        Returns:
            A boolean to determine if it is okay to assign this employee to 
            desired schedule. Returns true always if employee has no warning
            flag in string.
        """
        
        pattern = r'\(\w\)'
        warning_flag = re.search(pattern, employee_str)
            
        if warning_flag:
            key = warning_flag.group(0)
            msg = tkMessageBox.askquestion("Potential Conflict With Selected Employee", 
                                       self.ELIGABLE_WARNING[key],
                                       icon='warning')
            if msg == 'yes':
                return True
            else:
                return False
                
        return True
    
    
    def display_eligables(self):
        """Load str of employee names into listbox and click assigned employee.
        
        This method loads the sorted eligable list as a list of strings. If the
        schedule has an assigned employee the corresponding string with that 
        employees name will be clicked.
        """
        
        employee_names = self.e_model.get_eligables()
        for e_name in employee_names:
            self.eligable_listbox.insert(tk.END, e_name)
            
        self.click_assigned_employee()
            
            
    def click_assigned_employee(self):
        """Set listbox curselect to employee assigned to schedule, if any."""
        self.eligable_listbox.selection_clear(0, tk.END)
        index = self.e_model.get_assigned_employee()
        if index != -1:
            self.eligable_listbox.selection_set(index)
        
            
    def destroy(self):
        """Destroy eligable view widgets."""
        self.eligable_listbox.destroy()
            
        
        
class EligableModel(object):
    """Represent the sorted eligable employees list for a particular schedule.
    
    The eligable model is essentially a sorted list of employee primary keys,
    then a dictionary to map the primary keys to a string that will be
    displayed in the view.
    
    Attributes:
        schedule_pk: primary key for corresponding schedule.
        dep: department of the current displayed calendar.
        day_model: model representing the day for given date/department.
        eligable_id_list: sorted list of employee id numbers.
    """

    def __init__(self, session, schedule_pk, department, day_model):
        """Initialize the model of particular day and department
    
        day_model uses the supplied arguments date and dep to fetch appropriate
        database schedules then processes their fields to create a model
        representation of that day.
        
        Args:
            calendar_display: calendar containing reference when updating costs.
            date: datetime.date object representing date for this day.
            week_number: int of week in the month.
            weekday: int of day in the month.
            dep: string representing the department this day represents.
        """
    
        self.session = session
        self.schedule_pk = schedule_pk
        self.dep = department
        self.day_model = day_model
        
        self.eligable_id_list = []
        
    
    def get_eligables(self):
        """Get a sorted str list of employee names and parallel list of id's
        
        This algorithm executes a multi-stage filter then sort process. The 
        top down explanation is as follows:
            1) Query all employees in database.
            
            2) Filter out all employees who are not assigned to department of
                the corresponding schedule's department.
                
            3) For each employee, given the schedule, use employee method to
                determine what 'tier' of availability they have. The tier of
                availability is marked with the following flags:
                    (A) - Available: No scheduling conflicts, overtime,
                          vacations, or repeating unavailability.
                    (O) - Overtime: Employee is available but currently is 
                          working enough hours that week to qualify for 
                          overtime.
                    (U) - Unavailable-Repeat: Employee has a repeat 
                          unavailability overlapping with the schedule.
                    (V) - Vacation: Employee has a vacation time period 
                          overlapping with the schedule.
                    (S) - Schedule-conflict: Employee has another schedule they 
                          are working on that overlaps with the schedule.
                
                With the returned tiered availability marker, use returned flag
                as a key to assign employee to that key in sorted dictionary.
                
           
            4) For each key in the sorted dictionary:

                a) Sort all employees in each key by how many hours they are 
                   currently assigned for this month, lowest amount of hours at 
                   beginning of list, highest amount of hours at end of list.
                
                b) Reverse the list of employees in the current given key, then
                   for each employee in list, check if their primary working 
                   department is the department of the schedule. If they are,
                   move them to the front of the list. (By doing this in
                   reverse order we preseve the sorted by hours work property
                   on top of sorting by primary department. For example, the
                   employee who has a primary department corresponding with the
                   schedule and the least amount of assigned hours will be at 
                   the beginning of the list.)
                   
                c) For each employee in this key, append a string of the 
                   employee name to the e_listbox_list array. Also, if the 
                   availability flag is any flag but (A), add this flag to the 
                   string to be appended. This list will be used for
                   the view.
                   
                d) For each employee in this key, append the database id to a
                   parallel list called employee_list. This parallel list will 
                   be used for the model.
                   
            5) Set employee_list as eligable_id_list in this instance, then 
               return e_listbox_list for use in the view.
        
        Returns:
            A list of strings representing the eligable employees in sorted 
            order according to their 'availability' for use in the view.
        """
        
        eligables = collections.OrderedDict([('(A)', []), ('(O)', []), ('(U)', []), ('(V)', []), ('(S)', [])])
        employee_list = []
        e_listbox_list = []
        employees = self.session.query(Employees).all()
        employees = [e for e in employees if (e.primary_department == self.dep
                                              or e.alternate1_department == self.dep
                                              or e.alternate2_department == self.dep)]
        db_schedule = self.get_db_schedule(self.schedule_pk)
        for e in employees:
            availability = e.get_availability(db_schedule)
            eligables[availability].append(e)
        # Sort in terms of scheduled hours, least hours at start of list
        # Then place employees with primary department at start of list 
        for key, e_list in eligables.iteritems():
            e_list.sort(key=lambda e: e.scheduled_hours)
            # We reverse the list for sorting accordin primary departments
            # because as we re-insert primary people it maintains the sorted
            # property with respect to primary department employees first,
            # sub-sorted with respect to scheduled hours
            for e in reversed(e_list):
                if e.primary_department == self.dep:
                    e_index = e_list.index(e)
                    e_list.insert(0, e_list.pop(e_index))
            for e in e_list:
                if key == '(A)':
                    e_listbox_list.append(e.first_name)
                else:
                    e_listbox_list.append(key + " " + e.first_name)

            sorted_employee_ids = [e.employee_id for e in e_list]
            employee_list += sorted_employee_ids
        self.eligable_id_list = employee_list
        return e_listbox_list
        
        
    def assign_employee_schedule(self, index):
        """Assign employee to schedule given list index.
        
        This method is inputted an index from a parallel list. (likely from
        some sort of listbox curselection.) In the case that the index
        coincides with the employee id already assigned to the schedule, does
        nothing. This method uses the parallel list of employee id's to find
        the corresponding id to the name clicked by user than assigns that 
        to the schedule.
        
        Args:
            index: The index in the sorted list of employees
        Returns:
            A string to represent the new string representing the schedule
            given the employee to be assigned to schedule. Returns None if the
            employee to be assigned is already assigned (nothing to change).
        """
        
        new_employee_id = self.eligable_id_list[index]
        new_employee = self.get_db_employee(new_employee_id)
        db_schedule = self.get_db_schedule(self.schedule_pk)
        old_employee_id = db_schedule.employee_id
        old_employee = self.get_db_employee(old_employee_id)
        if old_employee and new_employee_id != old_employee_id:
            old_employee.remove_schedule(db_schedule)
            db_schedule.employee_id = new_employee.employee_id
            new_employee.add_schedule(db_schedule)
            self.session.commit()
                
            self.day_model.update_costs()
            new_schedule_str = self.day_model.get_schedule_str(db_schedule)
            return new_schedule_str
        elif db_schedule.employee_id == None: 
            db_schedule.employee_id = new_employee.employee_id
            new_employee.add_schedule(db_schedule)
            self.session.commit()
            
            self.day_model.update_costs()
            new_schedule_str = self.day_model.get_schedule_str(db_schedule)
            return new_schedule_str
        # Case where employee to be assigned is already assigned
        else:
            return None
        
        

    def get_assigned_employee(self):
        """Get index of employee assigned to schedule in sorted eligables list.
        
        Given the sorted list of eligable employees for a given schedule this
        method tries to find the employee assigned to it in the list of 
        eligable employees. In the case where no employee has been assigned
        to this schedule the method returns -1 to indicate no employee has been
        assigned. Note that if an employee is assigned to a schedule then that
        employee will always be in the list of eligable employees.
        
        Returns:
            An integer representing the index of the assigned employee of the
            schedule with respect to the sorted list of eligable employees.
            Returns -1 if no employee is assigned to the schedule.
        """
        db_schedule = self.get_db_schedule(self.schedule_pk)
        if db_schedule.employee_id:
            employee = (self.session
                            .query(Employees)
                            .filter(Employees.employee_id 
                                    == db_schedule.employee_id)
                            .first())
                            
            index = self.eligable_id_list.index(employee.employee_id)
            return index
        else:
            return -1
            
            
    def get_db_schedule(self, id):
        """Get the database schedule given its primary key.
        
        Args:
            id: primary key of the schedule to query from database.
        
        Returns:
            The database schedule object corresponding to supplied primary key.
        """
        
        db_schedule = (self.session.query(DB_Schedule)
                                   .filter(DB_Schedule.id == id)
                                   .first())
        return db_schedule
        
        
    def get_db_employee(self, id):
        """Get the database employee given its primary key.
        
        Args:
            id: primary key of the employee to query from database.
        
        Returns:
            The database employee object corresponding to supplied primary key.
        """
        
        employee = (self.session
                        .query(Employees)
                        .filter(Employees.employee_id == id)
                        .first())    
        return employee



class ScheduleWidget(tk.Frame):
    """Composite widget to represent a clickable schedule
    
    The ScheduleWidget allows for user to click on a tk.Label that highlights
    color on both mouseover and mouse clicks. There is a remove button the user
    can click which will then remove the schedule from the view and the 
    database altogether.
    
    Attributes:
        parent: tk.Frame container for the child widgets.   
        pk: Primary key for the schedule for reference.
        day_vc: View and controller reference for the date schedule is assigned.
    """
    
    def __init__(self, parent, pk, schedule_str, day_vc):
        """Initialize clickable schedule representation.

        Args:
            parent: tk.Frame container for the child widgets.   
            pk: Primary key for the schedule for reference.
            schedule_str: String representation of the schedule for view
                display.
            day_vc: View and controller reference for the date schedule is 
                assigned.
        """
    
        self.parent = parent
        self.pk = pk
        self.day_vc = day_vc
          
        self.schedule_frame = tk.Frame(self.parent, bg="white", 
                                       borderwidth=1, relief=tk.RIDGE)
        self.str_var = tk.StringVar(self.schedule_frame)
        self.str_var.set(schedule_str)
        self.schedule_label = tk.Label(self.schedule_frame, 
                                       width=20, height=1, 
                                       anchor=tk.NW, 
                                       font=('Tahoma', 11, tk.NORMAL),
                                       textvariable=self.str_var, 
                                       bg="white")                   
        self.schedule_label.bind('<Enter>', lambda event: self.mouse_enter(event))
        self.schedule_label.bind('<Leave>', lambda event: self.mouse_leave(event))
        self.remove_button = ttk.Button(self.schedule_frame, text="Remove")
        # We use the rowid as a unique identifier for location
        # of index for clicking a schedule or remove button
        self.schedule_frame.bind("<Button-1>", 
                                 lambda event: self.day_vc.schedule_widget_click(event, self))
        self.schedule_label.bind("<Button-1>", 
                                 lambda event: self.day_vc.schedule_widget_click(event, self))
        self.remove_button.configure(command=lambda: self.remove())
        self.schedule_frame.pack()
        self.schedule_label.pack(side=tk.LEFT)
        self.remove_button.pack(side=tk.LEFT)
        
               
    def set_to_clicked(self):
        """Set the ScheduleWidget to be highlighted."""
        self.schedule_frame.config(bg="LightSkyBlue")
        self.schedule_label.config(bg="LightSkyBlue")
            

    def set_to_unclicked(self):
        """Set the ScheduleWidget to be un-highlighted."""
        self.schedule_frame.config(bg="white")
        self.schedule_label.config(bg="white")

        
    def mouse_enter(self, event):
        """Highlights schedule_widgets to display mouse enter event.
        
        Args:
            event: tk mouse over event object.
        """
        
        if self.day_vc.current_clicked_schedule != self.pk:
            self.schedule_frame.config(bg="#ccebff")
            self.schedule_label.config(bg="#ccebff")
        
        
    def mouse_leave(self, event):
        """De-highlights schedule_widgets to display mouse leave event.
        
        Args:
            event: tk mouse leave event object.
        """
        
        if self.day_vc.current_clicked_schedule != self.pk:
            self.schedule_frame.config(bg="white")
            self.schedule_label.config(bg="white")
            
            
    def set_text(self, str):
        """Sets the StringVar value of this schedule widget.
        
        Args:
            str: String that will set the new string to be displayed.
        """
        
        self.str_var.set(str)
        	
            
    def remove(self):
        """Call day_vc to remove schedule, then destroy self widgets."""
        self.day_vc.remove_schedule(self.pk)
        self.destroy()
        
    
    def destroy(self):
        """Destroy all the widgets of this ScheduleWidget."""
        self.schedule_frame.destroy()
        self.schedule_label.destroy()
        self.remove_button.destroy()
            
    


class ScheduleEditor(tk.Frame):
    """Composite widget to add schedules
    
    The ScheduleEditor allows for the user to add a schedule with several 
    choices: The hour and minute and whether to hide or show start/end times to
    the user. Due to the way the user interface is set up, the date and 
    department have already been selected, so the user only needs to specify 
    time and to hide or show the start/end times.
    
    Attributes:
        master: tk.Frame container for the child widgets.   
        c_page: Quasi-controller object for ScheduleEditor to
            know calendar to add schedules to.
    """

    def __init__(self, master, calendar_page):
        """Initialize interface for user to add schedules to calendar.

        Args:
            master: tk.Frame container for the child widgets.   
            calendar_page: Quasi-controller object for ScheduleEditor to
                know calendar to add schedules to.
        """
        
        self.c_page = calendar_page
        self.master = master
        
        # LabelFrame container for all child subwidgets
        self.schedule_frame = ttk.LabelFrame(self.master, text="Schedule Editor")
        self.schedule_frame.pack(fill="both")
        self.schedule_add_frame = tk.Frame(self.schedule_frame)
        self.schedule_add_frame.pack()
        self.add_schedule_button = ttk.Button(self.schedule_add_frame, 
                                             text='Add Schedule', 
                                             command=self.add_schedule)
        self.add_schedule_button.grid(row=8, column=0, columnspan=5, pady=7)
        
        # Start time widgets
        self.start_label = tk.Label(self.schedule_add_frame, 
                                    text = "Start time: ")
        self.start_label.grid(row=4, column=0)
        self.start_te = TimeEntry(self.schedule_add_frame)
        self.start_te.grid(row=4, column=1, columnspan=3)
        self.s_hide = tk.BooleanVar(self.schedule_add_frame)
        self.s_hide_cb = ttk.Checkbutton(self.schedule_add_frame, 
                                        onvalue=True, 
                                        offvalue=False, 
                                        variable=self.s_hide, 
                                        text="Hide Start Time")
        self.s_hide_cb.grid(row=5, column=1, columnspan=3, sticky=tk.W+tk.N)

        # End time widgets
        self.end_label = tk.Label(self.schedule_add_frame, text = "End time: ")
        self.end_label.grid(row=6, column=0)
        self.end_te = TimeEntry(self.schedule_add_frame)
        self.end_te.grid(row=6, column=1, columnspan=3)
        self.e_hide = tk.BooleanVar(self.schedule_add_frame)
        self.e_hide_cb = ttk.Checkbutton(self.schedule_add_frame, 
                                        onvalue=True, 
                                        offvalue=False, 
                                        variable=self.e_hide, 
                                        text="Hide End Time")
        self.e_hide_cb.grid(row=7, column=1, columnspan=3, sticky=tk.W+tk.N)
        
        # Frame to display interactive, clickable schedules for given day.
        self.schedule_widgets_frame = tk.Frame(self.schedule_frame)
        self.schedule_widgets_frame.pack(pady=6)


    def add_schedule(self):
        """Add a schedule to corresponding calendar currently loaded.
        
        Add schedule uses the current loaded calendar to determine what date
        is currently selected by the user, along with the department of the
        of the calendar, then combines this information with the start and 
        end times selected by the user along with hide/show booleans selected 
        by user to create a schedule object that will be added to a database 
        and view.
        """

        cal = self.c_page.calendar_display
        day_vc = cal.current_clicked_day
        day_model = day_vc.day_model
        schedule_date = datetime.date(cal.date.year, cal.date.month, 
                                      int(day_model.date.day))
        cal_date = datetime.date(cal.date.year, cal.date.month, 1)

        start_time = self.start_te.get()
        end_time = self.end_te.get()
        start_time_str = (str(cal.date.year) + " " 
                          + str(cal.date.month) + " " 
                          + str(day_model.date.day) +  " " 
                          + start_time[0] + " " 
                          + start_time[1] + " " 
                          + start_time[2])
        end_time_str =  (str(cal.date.year) + " " 
                         + str(cal.date.month) + " " 
                         + str(day_model.date.day) +  " " 
                         + end_time[0] + " " 
                         + end_time[1] + " " 
                         + end_time[2])
        start_datetime = datetime.datetime.strptime(start_time_str, 
                                                "%Y %m %d %I %M %p")
        end_datetime = datetime.datetime.strptime(end_time_str, 
                                              "%Y %m %d %I %M %p")
        # Ensure a valid begining and end time have been selected
        if start_datetime < end_datetime:
            id = day_model.insert_new_schedule(start_datetime, 
                                               end_datetime,
                                               self.s_hide.get(),
                                               self.e_hide.get(),
                                               cal.dep)
            day_vc.highlight_new_schedule(id)
        else:
            print("Invalid Schedule time. " 
                  "Start time begins after end time. Beep. Boop. Bop.")
            
   
            
class CalendarCalculator(tk.Frame):
    """Composite widget to display costs of all calendars for given month.
    
    The CalendarCalculator is essentially a list of percentages of employment
    cost for a given month's schedules relative to average monthly revenue
    for that given month. Using the current selected calendar's month and year,
    CalendarCalculator fetches this employment cost to revenue for all
    departments and displays them in a list. Lastly, CalendarCalculator then
    creates a summation of all percentages and uses this to display total cost
    of all employment for that month relative to average revenue for that
    month.
    
    Attributes:
        master: tk.Frame container for the child widgets.   
        cal: Quasi-controller object for ScheduleEditor to
            know calendar to add schedules to.
        percentage_dict: dictionary object that maps strings of department
            names to a tk.StringVar object that represent the percentage of 
            employment cost to average revenue for the current selected month 
            and year.
    """
    
    def __init__(self, master, session, calendar_display, dep_list):
        """Initialize the display of costs for calendars.

        Args:
            master: tk.Frame container for the child widgets.   
            calendar_page: Quasi-controller object for ScheduleEditor to
                know calendar to add schedules to.
            dep_list: A string list of all departments.
        """
    
        self.master = master
        self.session = session
        self.cal = calendar_display
        self.percentage_dict = {}
        
        
        calc_frame = ttk.LabelFrame(self.master, 
                                    text='Payroll To Revenue Ratio*')
        calc_frame.pack(fill=tk.X, pady=12)
        warning_lbl = tk.Label(calc_frame, 
                               text='*Alpha Vesion Does Not Include Benefits Cost')
        warning_lbl.pack()
        
        for k in dep_list:
            
            frame = tk.Frame(calc_frame, borderwidth=1)
            frame.pack()
            title = tk.Label(frame, 
                                  width=10, height=1,
                                  text=k)
            title.pack(side=tk.LEFT)
            percentage = tk.StringVar(frame)
            self.percentage_dict[k] = percentage
            percentage_label = tk.Label(frame, height=1,
                                        textvariable=percentage)
            percentage_label.pack(side=tk.LEFT)
            
        # Create a special calculator that is the total of all calendars
        frame = tk.Frame(calc_frame, borderwidth=1)
        frame.pack()
        title = tk.Label(frame, 
                              width=10, height=1,
                              text='Total')
        title.pack(side=tk.LEFT)  
        percentage = tk.StringVar(frame)
        self.percentage_dict['Total'] = percentage
        percentage_label = tk.Label(frame, height=1,
                                    textvariable=percentage)
        percentage_label.pack(side=tk.LEFT)
        
        self.update_costs()
        
        
    def get_percentage(self, schedules, total):
        """Return percentage cost of schedules relative to total."""
        sum = 0.0
        for s in schedules:
            sum += s.cost()
        percent = int(round((sum / total) * 100, 0))
        
        return percent

        
    def get_monthly_total_avg(self):
        """Return average total revenue for month and year."""
        monthly_sales = self.session.query(MonthSales).all()
        monthly_sales = [s for s in monthly_sales if (s.month_and_year.month
                                                      == self.cal.date.month)]
        
        average = 1
        for month_sales in monthly_sales:
            average += month_sales.total_sales
        if len(monthly_sales) != 0:
            average = average / len(monthly_sales)
            return average
        else:
            return None
        

    def update_costs(self):
        """Update list of all percentages."""
        monthly_avg = self.get_monthly_total_avg()
        # Case where there is no revenue data to compare schedule cost with
        if monthly_avg is None:
            for k in self.percentage_dict:
                var = self.percentage_dict[k]
                var.set('No Data')
            return
    
        schedules = (self.session
                         .query(DB_Schedule)
                         .filter(DB_Schedule.calendar_date == self.cal.date)
                         .all())
        departments = (self.session
                           .query(Department)
                           .all())
        dep_schedules = {}
        for d in departments:
            dep_schedules[d.name] = 0
        # Split up schedules into sublist according to assigned department
        for k in dep_schedules:
            schedule_coll = [s for s in schedules if s.department == k]
            dep_schedules[k] = schedule_coll
        
        # We find individual percentage costs for department and keep a
        # running sum in order to also calculate total percentage cost of all
        # departments.
        total = 0
        for k in dep_schedules:
            percent = self.get_percentage(dep_schedules[k], monthly_avg)
            total += percent
            var = self.percentage_dict[k]
            var.set((str(percent) + "%"))
        total_var = self.percentage_dict['Total']
        total_var.set((str(total) + "%"))
       
            
        
            
    
    
        
        

        